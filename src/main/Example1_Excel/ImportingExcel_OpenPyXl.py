import openpyxl


def main():
    workbook = openpyxl.load_workbook("../Names.xlsx")
    sheet = workbook["Sheet1"]
    for row in range(1, sheet.max_row + 1, 1):
        print(sheet.cell(row=row, column=sheet.max_column).value)


if __name__ == '__main__':
    main()
